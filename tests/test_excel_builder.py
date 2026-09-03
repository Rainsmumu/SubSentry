import unittest
from io import BytesIO

import openpyxl

from circuit_analyzer import _NORMALIZE_TYPE, _VALID_TYPES
from excel_builder import build_excel


class ExcelBuilderTests(unittest.TestCase):
    def test_additional_types_are_included_as_customer_circuits(self):
        self.assertIn("MPLS-VPN", _VALID_TYPES)
        self.assertEqual(_NORMALIZE_TYPE["MPLS-VPN"], "IEPL")
        self.assertIn("NNI", _VALID_TYPES)
        self.assertEqual(_NORMALIZE_TYPE["NNI"], "IEPL")

    def test_export_contains_routes_type_and_cooperation(self):
        circuit = {
            "customer": "测试客户",
            "circuit_id": "SEL/DAC-BEI/CU NP001",
            "bandwidth": "64K",
            "site_a": "上海",
            "site_b": "韩国",
            "impact_status": "无保护",
            "type": "IEPL",
            "source_type": "DDN",
            "cooperation": "Outbound",
            "route1": "APCN2 S4A",
            "route2": "NCP S3",
            "route3": "",
            "route4": "",
            "is_broken": True,
            "is_shanghai": True,
        }

        data = build_excel("APCN2_S4A", "2026-07-25T10:30", [circuit])
        workbook = openpyxl.load_workbook(
            BytesIO(data), read_only=True, data_only=True
        )
        sheet = workbook.active

        self.assertEqual(sheet.max_column, 14)
        self.assertEqual(sheet.cell(3, 9).value, "电路性质")
        self.assertEqual(sheet.cell(3, 10).value, "合作方式\n（Inbound/Outbound）")
        self.assertEqual(
            [sheet.cell(3, col).value for col in range(11, 15)],
            ["第一路由", "第二路由", "第三路由", "第四路由"],
        )
        self.assertEqual(sheet.cell(4, 9).value, "DDN")
        self.assertEqual(sheet.cell(4, 10).value, "Outbound")
        self.assertEqual(sheet.cell(4, 11).value, "APCN2 S4A")
        self.assertEqual(sheet.cell(4, 12).value, "NCP S3")

        workbook.close()


if __name__ == "__main__":
    unittest.main()
