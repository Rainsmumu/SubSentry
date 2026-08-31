import unittest
from io import BytesIO

import openpyxl

from supervisor_excel import build_supervisor_excel


class SupervisorExcelTests(unittest.TestCase):
    def test_stage_columns_and_changed_content_are_red(self):
        event = {
            "cable_id": "TPE_S4",
            "fault_time": "2026-08-30T14:37",
            "workflow": {
                "fields": {
                    "first_report": {"complaint_count": "1"},
                    "breakpoint_report": {"distance_km": "126.4", "complaint_count": "2"},
                    "repair_plan_report": {
                        "repair_start_date": "2026-09-02",
                        "expected_restore_date": "2026-09-04",
                        "complaint_count": "3",
                    },
                    "final_report": {
                        "actual_restore_time": "2026-09-03T22:18",
                        "fault_cause": "经TPE NOC确认，海缆受到外力损伤。",
                        "complaint_count": "3",
                    },
                },
            },
        }
        summary = {
            "ip_loss_gbps": 13,
            "ip_loss_sh_gbps": 13,
            "iepl_broken": 4,
            "iepl_broken_sh": 3,
            "iepl_broken_gbps": 130,
            "iepl_broken_sh_gbps": 120,
        }

        data = build_supervisor_excel(event, summary, "final_report")
        workbook = openpyxl.load_workbook(BytesIO(data), rich_text=True)
        sheet = workbook.active

        self.assertEqual(sheet.max_column, 5)
        self.assertEqual(sheet.cell(1, 5).value, "上海国际局（第三次续报）")
        self.assertIn("用户申告1起", sheet.cell(11, 2).value)
        self.assertEqual(sheet.cell(6, 2).value, sheet.cell(6, 5).value)
        # 故障地点在第一次续报发生变化，整格红色。
        self.assertEqual(sheet.cell(8, 3).font.color.rgb, "00FF0000")
        # 故障名称全程不变，保持黑色。
        self.assertNotEqual(getattr(sheet.cell(6, 5).font.color, "rgb", None), "00FF0000")
        self.assertIn("均已恢复", str(sheet.cell(10, 5).value))
        workbook.close()


if __name__ == "__main__":
    unittest.main()
