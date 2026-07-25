"""
data_source.py — 槽路表数据源管理

统一管理"当前使用的槽路表 Excel 文件"：
  - 优先使用用户通过网页上传、保存到 data/uploads/ 的最新文件；
  - 若尚未上传，则回退到项目内自带的样例文件
    "上海ITMC电路槽路表0407改进版.xlsx"。

只读取其中的 "金桥机房电路" sheet。上传时会做基础校验
（扩展名、能否打开、是否包含目标 sheet），并记录上传时间等元信息。
"""

from __future__ import annotations

import json
import os
import shutil
import tempfile
from datetime import datetime

import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.abspath(
    os.environ.get("SUBSENTRY_DATA_DIR", os.path.join(BASE_DIR, "data"))
)
BACKUP_DIR = os.path.abspath(
    os.environ.get("SUBSENTRY_BACKUP_DIR", os.path.join(DATA_DIR, "backups"))
)

# 需要读取的目标 sheet 名
TARGET_SHEET = "金桥机房电路"

# 上传目录与当前数据源固定文件名
UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")
CURRENT_FILE = os.path.join(UPLOAD_DIR, "current_circuit_table.xlsx")
META_FILE = os.path.join(UPLOAD_DIR, "current_meta.json")

# 尚未上传时使用的初始数据源（项目自带样例）
DEFAULT_SOURCE = os.path.abspath(
    os.environ.get(
        "SUBSENTRY_DEFAULT_SOURCE",
        os.path.join(BASE_DIR, "上海ITMC电路槽路表0407改进版.xlsx"),
    )
)
# 旧版固定文件（作为再兜底，兼容历史部署）
LEGACY_SOURCE = os.path.join(BASE_DIR, "金桥机房电路表.xlsx")

ALLOWED_EXT = {".xlsx", ".xlsm"}


class DataSourceError(Exception):
    """数据源校验/保存过程中的可读错误。"""


def get_current_path() -> str:
    """返回当前应读取的数据源文件绝对路径（含回退逻辑）。"""
    if os.path.exists(CURRENT_FILE):
        return CURRENT_FILE
    if os.path.exists(DEFAULT_SOURCE):
        return DEFAULT_SOURCE
    return LEGACY_SOURCE


def _read_meta() -> dict:
    if os.path.exists(META_FILE):
        try:
            with open(META_FILE, encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            pass
    return {}


def _write_meta(meta: dict) -> None:
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(prefix="meta-", suffix=".json", dir=UPLOAD_DIR)
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp_path, META_FILE)
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def _backup_current_source() -> None:
    """上传新槽路表前备份当前文件和元信息，最多保留最近30份。"""
    if not os.path.exists(CURRENT_FILE):
        return

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
    target_dir = os.path.join(BACKUP_DIR, f"datasource-{timestamp}")
    os.makedirs(target_dir, exist_ok=True)
    shutil.copy2(CURRENT_FILE, os.path.join(target_dir, "current_circuit_table.xlsx"))
    if os.path.exists(META_FILE):
        shutil.copy2(META_FILE, os.path.join(target_dir, "current_meta.json"))

    backups = sorted(
        (
            os.path.join(BACKUP_DIR, name)
            for name in os.listdir(BACKUP_DIR)
            if name.startswith("datasource-")
            and os.path.isdir(os.path.join(BACKUP_DIR, name))
        ),
        key=os.path.getmtime,
        reverse=True,
    )
    for old_dir in backups[30:]:
        shutil.rmtree(old_dir, ignore_errors=True)


def validate_workbook(path: str) -> None:
    """
    校验一个 Excel 文件是否可用作数据源。
    失败时抛出 DataSourceError，附带可读原因。
    """
    ext = os.path.splitext(path)[1].lower()
    if ext not in ALLOWED_EXT:
        raise DataSourceError(f"文件不是 Excel（仅支持 {'/'.join(sorted(ALLOWED_EXT))}）")

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:  # openpyxl 会抛多种异常
        raise DataSourceError(f"无法打开 Excel 文件：{e}")

    try:
        if TARGET_SHEET not in wb.sheetnames:
            raise DataSourceError(
                f"缺少必需的 sheet「{TARGET_SHEET}」。当前包含：{', '.join(wb.sheetnames)}"
            )
        ws = wb[TARGET_SHEET]
        # 基础表结构检查：至少要有表头 + 数据，且列数覆盖到电路状态列
        if ws.max_row < 3 or ws.max_column < 91:
            raise DataSourceError("「金桥机房电路」表结构不符合预期（行/列数不足）")
    finally:
        wb.close()


def save_uploaded(tmp_path: str, original_name: str) -> dict:
    """
    校验并把上传的临时文件保存为当前数据源。
    返回最新的数据源元信息 dict。校验失败抛 DataSourceError。
    """
    validate_workbook(tmp_path)

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    _backup_current_source()
    fd, staged_path = tempfile.mkstemp(
        prefix="circuit-table-", suffix=".xlsx", dir=UPLOAD_DIR
    )
    os.close(fd)
    try:
        shutil.copyfile(tmp_path, staged_path)
        os.replace(staged_path, CURRENT_FILE)
    finally:
        if os.path.exists(staged_path):
            os.remove(staged_path)

    meta = {
        "original_name": original_name,
        "uploaded_at": datetime.now().isoformat(timespec="seconds"),
        "size_bytes": os.path.getsize(CURRENT_FILE),
    }
    _write_meta(meta)
    return get_status()


def get_status() -> dict:
    """
    返回当前数据源状态，供前端展示：
      - filename:    当前使用的文件展示名
      - is_uploaded: 是否为用户上传（否则为项目自带默认样例）
      - uploaded_at: 上传时间（仅上传时有）
      - path:        实际读取路径（相对项目根）
      - readable:    能否正常读取「金桥机房电路」sheet
      - circuit_count: 开通且属于 IP/IEPL/IPLC/DDN 的电路条数（读取成功时）
      - error:       读取失败时的原因
    """
    path = get_current_path()
    is_uploaded = os.path.exists(CURRENT_FILE)
    meta = _read_meta() if is_uploaded else {}

    if is_uploaded:
        filename = meta.get("original_name") or os.path.basename(CURRENT_FILE)
    else:
        filename = os.path.basename(path)

    try:
        display_path = (
            os.path.relpath(path, BASE_DIR)
            if os.path.commonpath([BASE_DIR, path]) == BASE_DIR
            else path
        )
    except ValueError:
        # Windows 不同盘符之间不能计算 commonpath/relpath。
        display_path = path

    status: dict = {
        "filename": filename,
        "is_uploaded": is_uploaded,
        "uploaded_at": meta.get("uploaded_at"),
        "path": display_path,
        "readable": False,
        "circuit_count": None,
        "error": None,
    }

    # 注意：不在此处调用 validate_workbook()——它会重新打开 13MB Excel，
    # 使每次打开页面都卡数秒。校验只在上传时(save_uploaded)做一次即可；
    # 状态读取直接用带缓存的 load_circuits()，数据源变化时会自动重载。
    try:
        from circuit_analyzer import load_circuits
        circuits = load_circuits()
        status["readable"] = True
        status["circuit_count"] = len(circuits)
    except FileNotFoundError:
        status["error"] = "找不到数据源文件"
    except KeyError:
        status["error"] = f"数据源缺少「{TARGET_SHEET}」sheet"
    except Exception as e:
        status["error"] = f"读取失败：{e}"

    return status
