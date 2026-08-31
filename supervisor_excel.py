"""生成崇明出口海缆的专业主管分阶段传报 Excel。"""

from __future__ import annotations

import io
from datetime import datetime

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from cable_config import CABLE_BY_ID
from circuit_analyzer import format_gbps
from fault_workflow import ensure_event_workflow


SUPERVISOR_STAGES = [
    ("first_report", "上海国际局（首报）"),
    ("breakpoint_report", "上海国际局（续报）"),
    ("repair_plan_report", "上海国际局（第二次续报）"),
    ("final_report", "上海国际局（第三次续报）"),
]

ROWS = [
    "上报专业", "子专业", "故障类别", "故障级别", "故障名称", "故障发生时间",
    "故障地点", "故障现象", "业务影响范围", "投诉情况", "故障原因", "故障处理责任人",
]


def _parse_time(value: str) -> datetime:
    for fmt in (
        "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"
    ):
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            pass
    raise ValueError(f"无法解析时间：{value}")


def _full_time(value: str) -> str:
    if not value:
        return "XXXX年XX月XX日"
    dt = _parse_time(value)
    if "T" in value or " " in value:
        return f"{dt.year}年{dt.month}月{dt.day}日{dt.hour:02d}:{dt.minute:02d}"
    return f"{dt.year}年{dt.month}月{dt.day}日"


def _complaints(value) -> int:
    try:
        return max(0, int(value))
    except (TypeError, ValueError):
        return 0


def _impact_scope(summary: dict) -> str:
    return (
        f"影响互联网带宽{format_gbps(summary['ip_loss_gbps'])}；"
        f"影响无保护及主备双断客户专线{summary['iepl_broken']}条，"
        f"共{format_gbps(summary.get('iepl_broken_gbps', 0))}，其中上海落地"
        f"{summary['iepl_broken_sh']}条、共"
        f"{format_gbps(summary.get('iepl_broken_sh_gbps', 0))}"
    )


def _stage_values(event: dict, summary: dict) -> dict[str, dict[str, str]]:
    ensure_event_workflow(event)
    cable = CABLE_BY_ID[event["cable_id"]]
    fields = event["workflow"]["fields"]
    fault_time = _parse_time(event["fault_time"])
    cable_label = f"{cable['cable']}海缆{cable['segment']}段"
    route = cable["route_desc"]
    landing = cable["landing"]
    responsible = cable["responsible"]
    fixed = {
        "上报专业": "国际网络",
        "子专业": "国际海陆缆",
        "故障类别": "国际海陆缆",
        "故障级别": "重大故障",
        "故障名称": f"{cable_label}上海{landing}出口故障",
        "故障发生时间": (
            f"{fault_time.year}年{fault_time.month}月{fault_time.day}日"
            f"{fault_time.hour:02d}:{fault_time.minute:02d}"
        ),
        "故障现象": f"{cable_label}（{route}）中断，业务受到影响。",
        "故障处理责任人": f"{responsible['name']} {responsible['phone']}",
    }

    pending_cause = "待电信核查确认。"
    first_fields = fields["first_report"]
    first = {
        **fixed,
        "故障地点": f"{cable['cable']} {cable['segment']}海缆NOC定位中",
        "业务影响范围": "业务影响范围正在统计中。",
        "投诉情况": (
            f"截至目前收到用户申告{_complaints(first_fields.get('complaint_count'))}起。"
        ),
        "故障原因": pending_cause,
    }

    breakpoint = fields["breakpoint_report"]
    breakpoint_location = (
        f"{cable_label}（{route}）中断，本次故障大致位于"
        f"{landing}外约{breakpoint.get('distance_km') or 'XX'}公里"
    )
    impact = _impact_scope(summary)
    breakpoint_values = {
        **fixed,
        "故障地点": breakpoint_location,
        "业务影响范围": impact,
        "投诉情况": (
            f"截至目前收到用户申告{_complaints(breakpoint.get('complaint_count'))}起。"
        ),
        "故障原因": pending_cause,
    }

    repair = fields["repair_plan_report"]
    repair_location = (
        f"{breakpoint_location}，计划于{_full_time(repair.get('repair_start_date', ''))}"
        f"开始维修，预计{_full_time(repair.get('expected_restore_date', ''))}修复"
    )
    repair_values = {
        **fixed,
        "故障地点": repair_location,
        "业务影响范围": impact,
        "投诉情况": f"截至目前收到用户申告{_complaints(repair.get('complaint_count'))}起。",
        "故障原因": pending_cause,
    }

    final = fields["final_report"]
    final_values = {
        **fixed,
        "故障地点": (
            f"{breakpoint_location}，已于{_full_time(final.get('actual_restore_time', ''))}修复"
        ),
        "业务影响范围": f"{impact}；目前受影响业务均已恢复。",
        "投诉情况": f"{_complaints(final.get('complaint_count'))}起用户申告已修复。",
        "故障原因": final.get("fault_cause") or pending_cause,
    }

    return {
        "first_report": first,
        "breakpoint_report": breakpoint_values,
        "repair_plan_report": repair_values,
        "final_report": final_values,
    }


def _set_changed_value(cell, current: str, previous: str | None) -> None:
    """继承文字保持黑色；追加文字标红；完全变化时整格标红。"""
    base_font = Font(name="宋体", size=10)
    red_font = Font(name="宋体", size=10, color="FF0000")
    if previous is None or current == previous:
        cell.value = current
        cell.font = base_font
        return
    if previous and current.startswith(previous):
        cell.value = CellRichText(
            TextBlock(InlineFont(rFont="宋体", sz=10), previous),
            TextBlock(InlineFont(rFont="宋体", sz=10, color="FF0000"), current[len(previous):]),
        )
        cell.font = base_font
        return
    cell.value = current
    cell.font = red_font


def build_supervisor_excel(
    event: dict, summary: dict, through_stage: str = "final_report"
) -> bytes:
    cable = CABLE_BY_ID[event["cable_id"]]
    if cable["landing"] != "崇明":
        raise ValueError("专业主管文件仅适用于崇明出口海缆")
    stage_ids = [stage_id for stage_id, _ in SUPERVISOR_STAGES]
    if through_stage not in stage_ids:
        raise ValueError("无效的专业主管文件阶段")
    included = SUPERVISOR_STAGES[:stage_ids.index(through_stage) + 1]
    values = _stage_values(event, summary)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "专业主管传报"
    sheet.freeze_panes = "B2"
    sheet.sheet_view.showGridLines = False

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

    sheet.column_dimensions["A"].width = 22
    sheet.cell(1, 1, "")
    for column, (_, title) in enumerate(included, 2):
        sheet.column_dimensions[get_column_letter(column)].width = 52
        cell = sheet.cell(1, column, title)
        cell.font = Font(name="宋体", size=10)
        cell.alignment = alignment
        cell.border = border
    sheet.cell(1, 1).border = border
    sheet.row_dimensions[1].height = 28

    previous_values = None
    for row_index, row_name in enumerate(ROWS, 2):
        label_cell = sheet.cell(row_index, 1, row_name)
        label_cell.font = Font(name="宋体", size=10)
        label_cell.alignment = alignment
        label_cell.border = border
        sheet.row_dimensions[row_index].height = 42

        previous_values = None
        for column, (stage_id, _) in enumerate(included, 2):
            current = values[stage_id][row_name]
            previous = previous_values[row_name] if previous_values else None
            cell = sheet.cell(row_index, column)
            _set_changed_value(cell, current, previous)
            cell.alignment = alignment
            cell.border = border
            previous_values = values[stage_id]

    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()
