"""
excel_builder.py — 国际海缆故障影响业务统计表生成

对齐样例文件格式：
  Row 1: 标题（合并 A1:H1，加粗居中）
  Row 2: 事件描述（合并 A2:H2）
  Row 3: 表头（客户专线电路 | 序号 | 客户名称 | 电路编号 | 电路速率 | A端 | Z端 | 此次故障影响主用或备用）
  Row 4+: IEPL 数据行，上海落地优先，按带宽从大到小排序
"""

from __future__ import annotations

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from cable_config import CABLE_BY_ID
from circuit_analyzer import circuit_sort_key


def _parse_fault_time(fault_time: str) -> datetime:
    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(fault_time.strip(), fmt)
        except ValueError:
            pass
    raise ValueError(f"无法解析故障时间：{fault_time}")


def build_excel(
    cable_id: str,
    fault_time: str,
    circuits: list[dict],
) -> bytes:
    """
    生成统计表 xlsx 并返回字节流。

    circuits: circuit_analyzer.analyze() 返回的 circuits 列表，
              只包含 type=="IEPL" 的电路。
    """
    cable = CABLE_BY_ID.get(cable_id)
    if not cable:
        raise ValueError(f"未知海缆 id：{cable_id}")

    dt = _parse_fault_time(fault_time)
    month  = dt.month
    day    = dt.day
    hour   = dt.hour
    minute = dt.minute

    cable_name = cable["cable"]
    segment    = cable["segment"]

    # 只统计 IEPL，按排序规则排列
    iepl = [c for c in circuits if c["type"] == "IEPL"]
    iepl.sort(key=circuit_sort_key)

    # ── 创建工作簿 ──────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ── 列宽 ────────────────────────────────────────────────────────
    col_widths = [14, 6, 28, 32, 10, 10, 10, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 通用样式 ────────────────────────────────────────────────────
    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_va  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_va    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    title_font  = Font(bold=True, size=14)
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill("solid", fgColor="D9E1F2")  # 浅蓝背景

    # ── Row 1：标题 ─────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 28
    cell = ws["A1"]
    cell.value     = "国际海缆故障影响业务统计表"
    cell.font      = title_font
    cell.alignment = center_va

    # ── Row 2：事件描述 ─────────────────────────────────────────────
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 22
    desc = (
        f"  {dt.year} 年 {month} 月 {day} 日 {hour} 时 {minute:02d} 分，"
        f" {cable_name}  海缆系统在  {segment}  段落故障。具体影响业务如下："
    )
    cell = ws["A2"]
    cell.value     = desc
    cell.alignment = left_va

    # ── Row 3：表头 ─────────────────────────────────────────────────
    headers = [
        "客户专线电路", "序号", "客户名称", "电路编号",
        "电路速率", "A端", "Z端",
        "此次故障影响主用或备用\n（如没有标注默认为无保护）",
    ]
    ws.row_dimensions[3].height = 36
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_va
        cell.border    = border_all

    # ── Row 4+：数据行 ───────────────────────────────────────────────
    for i, c in enumerate(iepl, 1):
        row = 3 + i
        ws.row_dimensions[row].height = 18

        # 如果是临时路由，在影响状态后添加标注
        impact_text = c["impact_status"]
        if c.get("is_temp_route", False):
            impact_text += " ⚠️临时路由"

        values = [
            None,                  # A: 客户专线电路（首列合并区，数据行留空）
            i,                     # B: 序号
            c["customer"],         # C: 客户名称
            c["circuit_id"],       # D: 电路编号
            c["bandwidth"],        # E: 电路速率
            c["site_a"],           # F: A端
            c["site_b"],           # G: Z端
            impact_text,           # H: 影响类型（含临时路由标注）
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = left_va
            cell.border    = border_all

    # ── 输出到字节流 ────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
