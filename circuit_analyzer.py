"""
circuit_analyzer.py — 电路影响分析引擎

从金桥机房电路表读取开通的 IP/IEPL/IPLC/DDN/MPLS-VPN 电路，
根据当前故障海缆集合判断每条电路的影响状态。

影响分类：
  无保护   — 单腿电路且该腿已断（业务中断）
  主备双断  — 多腿电路所有路由均已断（业务中断）
  主用     — 第一路由断但有存活备用（业务通过备用存活）
  备用     — 第一路由存活但某备用路由断（保护能力降低）

说明：路由字段来自槽路表的「第一路由/第二路由/第三路由/第四路由」四列，
均为实际路由腿。以前误把第四路由当作"当前路由(current_route)"并只检查它，
现已修正为把四条路由都作为并列路由腿处理。
"""

from __future__ import annotations

import os
import re
import openpyxl
from cable_config import match_route_to_cable_ids, CABLE_BY_ID

# 金桥机房电路表列索引（0-based）
# route1..route4 分别对应槽路表的「第一路由/第二路由/第三路由/第四路由」
_COL = {
    "site_a":      1,
    "site_b":      10,
    "customer":    12,
    "circuit_id":  22,
    "route1":      26,   # 第一路由（主用）
    "route2":      27,   # 第二路由
    "route3":      28,   # 第三路由
    "route4":      29,   # 第四路由（并列路由腿，非"当前路由"）
    "bandwidth":   34,
    "type":        38,
    "cooperation": 39,
    "status":      90,
}

# 统计的电路性质：IP，以及 IEPL/IPLC/DDN/MPLS-VPN 客户专线
_VALID_TYPES = {"IP", "IEPL", "IPLC", "DDN", "MPLS-VPN"}

# IPLC、DDN、MPLS-VPN 均按客户专线参与页面、通报及汇总；source_type 保留原始性质。
_NORMALIZE_TYPE = {"IPLC": "IEPL", "DDN": "IEPL", "MPLS-VPN": "IEPL"}

# 影响状态常量
STATUS_NO_PROTECT   = "无保护"      # 业务中断（单腿断）
STATUS_DUAL_BREAK   = "主备双断"    # 业务中断（多腿全断）
STATUS_PRIMARY      = "主用"        # 主用受影响（业务通过备用存活）
STATUS_BACKUP       = "备用"        # 备用受影响（主用存活）

# 中断状态集合（用于统计"中断"数量和带宽）
BROKEN_STATUSES = {STATUS_NO_PROTECT, STATUS_DUAL_BREAK}


# ── 数据加载（带缓存，数据文件不变时只读一次）──────────────────────
_data_cache: list | None = None
_data_mtime: float = 0.0
_data_path: str = ""
# 全量源表索引（不过滤状态/性质），用于对比诊断按国际电路名回查
_source_index_cache: dict | None = None
_source_index_key: tuple | None = None
_INTERNATIONAL_CIRCUIT_RE = re.compile(r"([A-Z0-9-]+(?:/[A-Z0-9-]+)+\s+[A-Z0-9-]+)", re.IGNORECASE)


def _get_excel_path() -> str:
    """当前使用的数据源路径（由 data_source 统一管理，支持网页上传）。"""
    from data_source import get_current_path
    return get_current_path()


def invalidate_cache() -> None:
    """清空电路数据缓存，使下次分析强制重新读取当前数据源。"""
    global _data_cache, _data_mtime, _data_path, _source_index_cache, _source_index_key
    _data_cache = None
    _data_mtime = 0.0
    _data_path = ""
    _source_index_cache = None
    _source_index_key = None


def _norm_intl(name: str) -> str:
    """归一化国际电路名作索引键：大写并去除所有空白。"""
    return "".join(str(name or "").upper().split())


def build_source_index(force_reload: bool = False) -> dict[str, list[dict]]:
    """
    按国际电路名索引槽路表「金桥机房电路」的**全部行**（不过滤状态/性质），
    供对比诊断在系统结果之外回查一条电路的真实状态、性质、路由，从而判定漏掉原因。

    返回 { 归一化国际电路名 -> [记录, ...] }，记录含 status/type/route1..4/site 等原始值。
    """
    global _source_index_cache, _source_index_key

    path = _get_excel_path()
    try:
        mtime = os.path.getmtime(path)
    except FileNotFoundError:
        raise FileNotFoundError(f"找不到槽路表数据源：{path}")

    key = (path, mtime)
    if not force_reload and _source_index_cache is not None and _source_index_key == key:
        return _source_index_cache

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["金桥机房电路"]

    index: dict[str, list[dict]] = {}
    for row in ws.iter_rows(min_row=3, max_col=95, values_only=True):
        raw = str(row[_COL["circuit_id"]] or "").strip()
        intl = extract_international_circuit_id(raw)
        nk = _norm_intl(intl)
        if not nk:
            continue
        index.setdefault(nk, []).append({
            "circuit_id":     intl,
            "circuit_id_raw": raw,
            "status":    str(row[_COL["status"]]    or "").strip(),
            "type":      str(row[_COL["type"]]      or "").strip(),  # 原始性质，未归一化
            "route1":    str(row[_COL["route1"]]    or "").strip(),
            "route2":    str(row[_COL["route2"]]    or "").strip(),
            "route3":    str(row[_COL["route3"]]    or "").strip(),
            "route4":    str(row[_COL["route4"]]    or "").strip(),
            "site_a":    str(row[_COL["site_a"]]    or "").strip(),
            "site_b":    str(row[_COL["site_b"]]    or "").strip(),
            "customer":  str(row[_COL["customer"]]  or "").strip(),
            "bandwidth": str(row[_COL["bandwidth"]] or "").strip(),
            "cooperation": str(row[_COL["cooperation"]] or "").strip(),
        })

    wb.close()
    _source_index_cache = index
    _source_index_key = key
    return index


def load_circuits(force_reload: bool = False) -> list[dict]:
    """
    加载金桥机房电路表中所有"开通"的 IP/IEPL/IPLC/DDN/MPLS-VPN 电路。
    返回列表，每项为字典形式的电路信息。
    """
    global _data_cache, _data_mtime, _data_path

    path = _get_excel_path()
    try:
        mtime = os.path.getmtime(path)
    except FileNotFoundError:
        raise FileNotFoundError(f"找不到槽路表数据源：{path}")

    # 数据源路径或修改时间变化时，缓存失效
    if (not force_reload and _data_cache is not None
            and mtime == _data_mtime and path == _data_path):
        return _data_cache

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["金桥机房电路"]

    circuits = []
    for row in ws.iter_rows(min_row=3, max_col=95, values_only=True):
        # 过滤：只取开通状态
        if str(row[_COL["status"]] or "").strip() != "开通":
            continue
        # 过滤：只取 IP / IEPL / IPLC / DDN / MPLS-VPN
        ctype = str(row[_COL["type"]] or "").strip()
        if ctype not in _VALID_TYPES:
            continue

        raw_circuit_id = str(row[_COL["circuit_id"]] or "").strip()

        circuits.append({
            "site_a":    str(row[_COL["site_a"]]    or "").strip(),
            "site_b":    str(row[_COL["site_b"]]    or "").strip(),
            "customer":  str(row[_COL["customer"]]  or "").strip(),
            "circuit_id_raw": raw_circuit_id,
            "circuit_id": extract_international_circuit_id(raw_circuit_id),
            "route1":    str(row[_COL["route1"]]    or "").strip(),
            "route2":    str(row[_COL["route2"]]    or "").strip(),
            "route3":    str(row[_COL["route3"]]    or "").strip(),
            "route4":    str(row[_COL["route4"]]    or "").strip(),
            "bandwidth": str(row[_COL["bandwidth"]] or "").strip(),
            "source_type": ctype,
            "type":      _NORMALIZE_TYPE.get(ctype, ctype),
            "cooperation": str(row[_COL["cooperation"]] or "").strip(),
        })

    wb.close()
    _data_cache = circuits
    _data_mtime = mtime
    _data_path = path
    return circuits


def extract_international_circuit_id(raw_value: str) -> str:
    """
    从"电路代号"列中提取国际电路编码。
    规则：
    1. 优先取形如 "SHI/CU-SEL/CU EP025" 的国际编码
    2. 忽略 BEARER、本地电路编码、纯中文说明
    3. 若未匹配到国际编码，则回退到第一条非空、非说明性文本
    """
    if not raw_value:
        return ""

    lines = [line.strip() for line in str(raw_value).splitlines() if line and line.strip()]
    for line in lines:
        match = _INTERNATIONAL_CIRCUIT_RE.search(line.upper())
        if match:
            return match.group(1).strip()

    for line in lines:
        upper = line.upper()
        if "BEARER" in upper or "本地电路编码" in line:
            continue
        return line
    return lines[0] if lines else ""


# ── 带宽解析 ────────────────────────────────────────────────────────

def bw_to_gbps(bw_str: str) -> float:
    """将带宽字符串转换为 Gbps 浮点值，用于汇总 IP 带宽损失。"""
    if not bw_str:
        return 0.0
    s = bw_str.strip().upper().replace(" ", "")
    try:
        if s.endswith("G"):
            return float(s[:-1])
        if s.endswith("M"):
            return float(s[:-1]) / 1000
        if s.endswith("K") or "K" in s:
            return float(s.replace("K", "")) / 1_000_000
        return float(s) / 1000  # 裸数字假设为 Mbps
    except (ValueError, TypeError):
        return 0.0


def format_gbps(gbps: float) -> str:
    """将 Gbps 浮点值格式化为通报用字符串，如 '128G'、'0.155G'。"""
    if gbps == int(gbps):
        return f"{int(gbps)}G"
    # 去掉末尾多余零，再加 G
    s = f"{gbps:.4f}".rstrip("0").rstrip(".")
    return s + "G"


# ── 核心分析逻辑 ────────────────────────────────────────────────────

def _route_legs(circuit: dict) -> list[str]:
    """返回该电路的所有非空路由腿（第一~第四路由）。"""
    return [r for r in (
        circuit.get("route1", ""),
        circuit.get("route2", ""),
        circuit.get("route3", ""),
        circuit.get("route4", ""),
    ) if r and r.strip()]


def _leg_is_broken(route_str: str, broken_ids: set[str]) -> bool:
    """
    判断一条路由腿是否受影响：
    该腿字段中包含的任一目标海缆段落命中了已断海缆集合即算断。
    支持拼接海缆（如 "APCN2 S4A+PC1崇明"）——只要包含 "APCN2 S4A" 即命中。
    """
    matched = match_route_to_cable_ids(route_str)
    return any(cid in broken_ids for cid in matched)


def _classify_circuit(circuit: dict, broken_ids: set[str]) -> str | None:
    """
    对单条电路进行影响分类。

    broken_ids: 当前已断海缆的 id 集合。
    路由腿取自第一~第四路由四列，均为并列路由腿；第一路由视为主用。

    返回影响状态字符串，或 None（不受影响，不纳入统计）。
    """
    routes = _route_legs(circuit)
    if not routes:
        return None

    is_broken = [_leg_is_broken(r, broken_ids) for r in routes]

    confirmed_broken = sum(is_broken)
    total = len(routes)

    # 完全不受影响
    if confirmed_broken == 0:
        return None

    # 全部路由腿断路
    if confirmed_broken == total:
        return STATUS_NO_PROTECT if total == 1 else STATUS_DUAL_BREAK

    # 第一路由（主用）断，尚有存活备用
    if is_broken[0]:
        return STATUS_PRIMARY

    # 第一路由存活，某备用路由断
    return STATUS_BACKUP


def circuit_sort_key(circuit: dict) -> tuple:
    """排序：真实中断优先，再按上海优先、带宽从大到小。"""
    broken_order = 0 if circuit.get("is_broken") else 1
    sh_order = 0 if circuit.get("is_shanghai") else 1
    return (
        broken_order,
        sh_order,
        -bw_to_gbps(circuit.get("bandwidth", "")),
        circuit.get("circuit_id", ""),
        circuit.get("customer", ""),
    )


def analyze(broken_cable_ids: list[str]) -> dict:
    """
    给定当前已断海缆 id 列表，分析所有受影响电路。

    返回结构：
    {
      "broken_ids": [...],
      "circuits": [
        {
          "site_a", "site_b", "customer", "circuit_id",
          "route1", "route2", "route3", "bandwidth", "type",
          "impact_status",    # 影响分类
          "is_shanghai",      # Site A == "上海"
          "is_broken",        # 业务是否中断（无保护 / 主备双断）
        },
        ...
      ],
      "summary": {
        "iepl_total": int,        # IEPL 受影响总条数（含主用/备用/疑似）
        "iepl_broken": int,       # IEPL 业务中断条数
        "iepl_broken_sh": int,    # 上海落地 IEPL 中断条数
        "ip_loss_gbps": float,    # IP 带宽损失（仅中断）
        "ip_loss_sh_gbps": float, # 上海落地 IP 带宽损失
      }
    }
    """
    broken_ids = set(broken_cable_ids)
    circuits = load_circuits()

    result_circuits = []
    for c in circuits:
        status = _classify_circuit(c, broken_ids)
        if status is None:
            continue
        is_sh = (c["site_a"] == "上海")
        is_broken = status in BROKEN_STATUSES
        result_circuits.append({
            **c,
            "impact_status": status,
            "is_shanghai": is_sh,
            "is_broken": is_broken,
        })

    result_circuits.sort(key=circuit_sort_key)

    # 汇总统计
    iepl_total     = sum(1 for c in result_circuits if c["type"] == "IEPL")
    iepl_broken    = sum(1 for c in result_circuits if c["type"] == "IEPL" and c["is_broken"])
    iepl_broken_sh = sum(1 for c in result_circuits if c["type"] == "IEPL" and c["is_broken"] and c["is_shanghai"])

    ip_loss     = sum(bw_to_gbps(c["bandwidth"]) for c in result_circuits if c["type"] == "IP" and c["is_broken"])
    ip_loss_sh  = sum(bw_to_gbps(c["bandwidth"]) for c in result_circuits
                      if c["type"] == "IP" and c["is_broken"] and c["is_shanghai"])

    return {
        "broken_ids": list(broken_ids),
        "circuits": result_circuits,
        "summary": {
            "iepl_total":      iepl_total,
            "iepl_broken":     iepl_broken,
            "iepl_broken_sh":  iepl_broken_sh,
            "ip_loss_gbps":    round(ip_loss, 4),
            "ip_loss_sh_gbps": round(ip_loss_sh, 4),
        },
    }
